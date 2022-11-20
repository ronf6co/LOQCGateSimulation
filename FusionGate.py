import os

import openpyxl
import xlsxwriter

SYMBOLIC = False

import sympy as sp
from sympy import *
import numpy as np
from photonic_circuit import PhotonicCircuit

if SYMBOLIC:
    Iden = sp.eye(2)
    X = sp.Matrix([[0, 1], [1, 0]])
    Y = sp.Matrix([[0, -1j], [1j, 0]])
    Z = sp.Matrix([[1, 0], [0, -1]])
else:
    I = np.eye(2)
    X = np.array([[0, 1], [1, 0]])
    Y = np.array([[0, -1j], [1j, 0]])
    Z = np.array([[1, 0], [0, -1]])


class FusionGate:

    def __init__(self, error_angle=0, error_axis=[1, 0, 0], angle_sign_list=[0,0,0,0,0]):
        if SYMBOLIC:

            self.error_axis = error_axis
            self.error_angle = error_angle

            self.R_error = sp.cos(error_angle / 2) * I + 1j * sp.sin(
                error_angle / 2) * \
                           (error_axis[0] * X + error_axis[0] * Y + error_axis[
                               0] * Z)
            X_sqrt = sp.cos(np.pi / 4) * I - sp.sin(np.pi / 4) * 1j * X

            circuit = PhotonicCircuit(4)
            circuit.apply_coupler(X_sqrt * self.R_error, 0, 1)
            circuit.apply_coupler(X_sqrt * self.R_error, 2, 3)
            circuit.apply_coupler(-1j * X * self.R_error, 1, 2)
            circuit.apply_coupler(X_sqrt * self.R_error, 0, 1)
            circuit.apply_coupler(X_sqrt * self.R_error, 2, 3)
            self.basis = [[0, 2], [0, 3], [1, 2], [1, 3]]
            self.U = circuit.project_on_computation_basis(self.basis)

        else:

            self.error_axis = error_axis
            self.error_angle = error_angle

            self.R_error = np.cos(error_angle / 2) * I + 1j * np.sin(
                error_angle / 2) * \
                           (error_axis[0] * X + error_axis[1] * Y + error_axis[
                               2] * Z)
            X_sqrt = np.cos(np.pi / 4) * I - np.sin(np.pi / 4) * 1j * X

            circuit = PhotonicCircuit(4)
            circuit.apply_coupler(
                np.matmul(X_sqrt, self.R_error_matrix(angle_sign_list[0])), 0, 1)
            circuit.apply_coupler(
                np.matmul(X_sqrt, self.R_error_matrix(angle_sign_list[1])), 2, 3)
            circuit.apply_coupler(
                -1j * np.matmul(X, self.R_error_matrix(angle_sign_list[2])), 1, 2)
            circuit.apply_coupler(
                np.matmul(X_sqrt, self.R_error_matrix(angle_sign_list[3])), 0, 1)
            circuit.apply_coupler(
                np.matmul(X_sqrt, self.R_error_matrix(angle_sign_list[4])), 2, 3)
            self.basis = [[0, 2], [0, 3], [1, 2], [1, 3]]
            self.U = circuit.project_on_computation_basis(self.basis)

    def R_error_matrix(self,sign):
        # sign = 0 -> + , sign = 1 -> -
        if SYMBOLIC:
            return sp.cos(self.error_angle / 2) * I + 1j * sp.sin(
                self.error_angle * (-1) ** sign / 2) * \
                   (self.error_axis[0] * X + self.error_axis[0] * Y +
                    self.error_axis[
                        0] * Z)
        else:
            return np.cos(self.error_angle / 2) * I + 1j * np.sin(
                self.error_angle * (-1) ** sign / 2) * \
                   (self.error_axis[0] * X + self.error_axis[1] * Y +
                    self.error_axis[
                        2] * Z)

    def getUFromClicks(self, clicks):
        if SYMBOLIC:
            return self.U.row(int(clicks, 2))
        else:
            return self.U[int(clicks, 2)]

    def print_latex_fusion_gate_state(self, clicks=""):
        # Symbols
        f1, f2, f3, f4 = symbols('f_1 f_2 f_3 f_4')
        a0, a1, b0, b1, c0, c1, d0, d1 = symbols('a_H^\dagger a_V^\dagger '
                                                 'b_H^\dagger b_V^\dagger '
                                                 'c_H^\dagger c_V^\dagger '
                                                 'd_H^\dagger d_V^\dagger')
        U = symbols('U')

        # U Matrix
        if clicks == "":
            U = Matrix(self.U)

            U
            simplify(U)

            # in out
            a_out = Matrix([[c0],
                            [c1],
                            [d0],
                            [d1], ])

            # Mul
            a_in = U * a_out
            a0, a1, b0, b1 = a_in[0], a_in[1], a_in[2], a_in[3]

            # state
            state = (f1 * a0 + f2 * a1) * (f3 * b0 + f4 * b1)

        else:
            U = Matrix((self.getUFromClicks(clicks)).transpose())
            a_out = Matrix([[f1 * f3],
                            [f1 * f4],
                            [f2 * f3],
                            [f2 * f4], ])
            state = U[0] * a_out[0] + U[1] * a_out[1] + U[2] * a_out[2] + U[3] * \
                    a_out[
                        3]

        return (simplify(
            collect(expand(state),
                    [c0 ** 2, c1 ** 2, d0 ** 2, d1 ** 2, c0 * d0, c0 * d1,
                     c1 * d0,
                     c0 * c1, d0 * d1])))







    def __str__(self):
        s = ""
        s += "\n______________________Fusion:__________________\n"
        s += "U : \n"
        if self.error_angle == 0 or SYMBOLIC:
            round_res = 7
        else:
            round_res = int(7+np.log10(np.abs(1/self.error_angle)))
        if not SYMBOLIC:
            s += str(np.round(self.U, round_res))
        else:
            s += str(sp.simplify(self.U))
        s += "\n\nBasis :\n"
        s += str(self.basis)

        s += "\n\nR error : \n"
        if not SYMBOLIC:
            s += str(np.round(self.R_error, round_res))
        else:
            s += str(self.R_error)
        s += "\n_________________________________________________"

        return s

    def __sub__(self, other):

        diff_mat = self.U-other.U
        return np.abs(np.sqrt(np.trace(np.matmul(diff_mat, np.conj(np.transpose(diff_mat))))))

    @staticmethod
    def findBestPMCombinationFusion(angle, error_axis):
        print("---------------------------------------")
        print("Finding combination")
        print("theta=", str(angle))
        print("Statrting...")
        c12naive = FusionGate(error_angle=angle, error_axis=error_axis,angle_sign_list=[0,0,0,0,0])
        c12i = FusionGate(error_angle=0, error_axis=error_axis,angle_sign_list=[0,0,0,0,0])
        worst_case = float("-Inf")
        best_case = float("Inf")
        for i in range(2 ** 5):
            b = format(i, '0' + str(5) + 'b')
            fusions_list = []
            for bit in b:
                if bit == '0':
                    fusions_list += [0]
                else:
                    fusions_list += [1]
            c12pm = FusionGate(error_angle=angle, error_axis=error_axis,angle_sign_list=fusions_list)
            print("for: "+str(b)+" PM dist from ideal : " + str(c12pm - c12i))
            if c12pm - c12i > worst_case:
                worst_case = c12pm - c12i
                worst_case_raw = b
            if c12pm - c12i < best_case:
                best_case = c12pm - c12i
                best_case_raw = b

        naive_case = c12naive - c12i
        print("Naive (00..0) dist from ideal : " + str(naive_case))

        print(
            "Worst case is : " + worst_case_raw + " value: " + str(worst_case))
        print("Best case is : " + best_case_raw + " value: " + str(best_case))
        print("---------------------------------------")

        # Add row to excel
        result_folder = r"C:\PycharmProjects\LOQCGateSimulation\outputs"
        result_file_name = r"C:\PycharmProjects\LOQCGateSimulation\outputs\1 fusion results.xlsx"
        if not (os.path.isfile(result_file_name) and os.access(result_file_name,
                                                               os.R_OK)):
            if not os.path.exists(result_folder):
                os.makedirs(result_folder)
            workbook = xlsxwriter.Workbook(result_file_name)
            worksheet = workbook.add_worksheet()
            workbook.close()

        book = openpyxl.load_workbook(result_file_name)

        sheet_name = "axis-" + str(error_axis[0]) + "," + str(
            error_axis[1]) + "," + str(error_axis[2]) + " theta-" + str(
            angle)
        if not sheet_name in book.sheetnames:
            ws = book.create_sheet(sheet_name)
        else:
            ws = book.get_sheet_by_name(sheet_name)

        data = {
                5: "Worst Combo", 6: "Best Combo", 7: "Naive Result",
                8: "Worst Result", 9: "Best Result",
                10: "Naive/Best", 11: "Worst/Best", 12: "Theta",
                13: "Angle Axis"}
        row_index = 1
        for col, value in data.items():
            ws.cell(row=row_index, column=col, value=value)
        book.save(filename=result_file_name)

        data = {
                5: worst_case_raw, 6: best_case_raw,
                7: naive_case, 8: worst_case, 9: best_case,
                10: naive_case / best_case, 11: worst_case / best_case,
                12: str(angle), 13: str(error_axis)}

        row_index = ws.max_row + 1
        ws.insert_rows(row_index)
        for col, value in data.items():
            ws.cell(row=row_index, column=col, value=value)

        book.save(filename=result_file_name)
#
# FusionGate.findBestPMCombinationFusion(0.2, [1,0,0])
# FusionGate.findBestPMCombinationFusion(0.2, [0,1,0])
# FusionGate.findBestPMCombinationFusion(0.2, [0,0,1])
