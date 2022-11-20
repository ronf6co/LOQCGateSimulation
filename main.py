import os

import openpyxl
import xlsxwriter

SYMBOLIC = False

from Cluster import Cluster
from FusionGate import FusionGate


def best_combo_with_vs_without_rotation(fi, fp, fm, cluster_size, clicks,
                                        fusions_amount, rot_state_angle,
                                        rot_state_axis=[1, 0, 0]):
    runOverPMCombinations(fi, fp, fm, cluster_size, clicks, fusions_amount,
                          rot_state_angle=0, rot_state_axis=rot_state_axis)
    runOverPMCombinations(fi, fp, fm, cluster_size, clicks, fusions_amount,
                          rot_state_angle=rot_state_angle,
                          rot_state_axis=rot_state_axis)


def runOverPMCombinations(fi, fp, fm, cluster_size, clicks, fusions_amount,
                          rot_state_angle=0, rot_state_axis=[1, 0, 0]):
    print("---------------------------------------")
    print("Finding combination for ", str(fusions_amount + 1), "cluster of",
          str(cluster_size), "  Clicks are ", clicks)
    print("State change from cluster is theta=", str(rot_state_angle))
    print("Statrting...")
    c12naive = Cluster.fuseLinearClusters(size_of_cluster=cluster_size,
                                          clicks=clicks,
                                          fusions=[fp] * fusions_amount,
                                          rot_state_angle=rot_state_angle,
                                          rot_state_axis=rot_state_axis)
    c12i = Cluster.fuseLinearClusters(size_of_cluster=cluster_size,
                                      clicks=clicks,
                                      fusions=[fi] * fusions_amount,
                                      rot_state_angle=rot_state_angle,
                                      rot_state_axis=rot_state_axis)
    worst_case = float("-Inf")
    best_case = float("Inf")
    for i in range(2 ** fusions_amount):
        b = format(i, '0' + str(fusions_amount) + 'b')
        fusions_list = []
        for bit in b:
            if bit == '0':
                fusions_list += [fp]
            else:
                fusions_list += [fm]
        c12pm = Cluster.fuseLinearClusters(size_of_cluster=cluster_size,
                                           clicks=clicks, fusions=fusions_list,
                                           rot_state_angle=rot_state_angle,
                                           rot_state_axis=rot_state_axis)
        if c12pm - c12i < 0.0178:
            print("for: "+str(b)+" PM dist from ideal : " + str(c12pm - c12i))
            print(
                "for: " + str(b) + " PM dist from ideal : " + str(c12pm - c12i))
        if c12pm - c12i > worst_case:
            worst_case = c12pm - c12i
            worst_case_raw = b
        if c12pm - c12i < best_case:
            best_case = c12pm - c12i
            best_case_raw = b

    naive_case = c12naive - c12i
    print("Naive (00..0) dist from ideal : " + str(naive_case))

    print("Worst case is : " + worst_case_raw + " value: " + str(worst_case))
    print("Best case is : " + best_case_raw + " value: " + str(best_case))
    print("**************************************************")

    # # Add row to excel
    # result_folder = r"C:\PycharmProjects\LOQCGateSimulation\outputs"
    # result_file_name = r"C:\PycharmProjects\LOQCGateSimulation\outputs\results.xlsx"
    # if not (os.path.isfile(result_file_name) and os.access(result_file_name,
    #                                                        os.R_OK)):
    #     if not os.path.exists(result_folder):
    #         os.makedirs(result_folder)
    #     workbook = xlsxwriter.Workbook(result_file_name)
    #     worksheet = workbook.add_worksheet()
    #     workbook.close()
    #
    # book = openpyxl.load_workbook(result_file_name)
    #
    # sheet_name = "axis-" + str(fp.error_axis[0]) + "," + str(
    #     fp.error_axis[1]) + "," + str(fp.error_axis[2]) + " theta-" + str(
    #     fp.error_angle)
    # if not sheet_name in book.sheetnames:
    #     ws = book.create_sheet(sheet_name)
    # else:
    #     ws = book.get_sheet_by_name(sheet_name)
    #
    # data = {2: "clicks", 3: "SubCluster Size", 4: "Cluster Amount",
    #         5: "Worst Combo", 6: "Best Combo", 7: "Naive Result",
    #         8: "Worst Result", 9: "Best Result",
    #         10: "Naive/Best", 11: "Worst/Best", 12: "Theta", 13: "Angle Axis"}
    # row_index = 1
    # for col, value in data.items():
    #     ws.cell(row=row_index, column=col, value=value)
    # book.save(filename=result_file_name)
    #
    # data = {2: clicks, 3: cluster_size, 4: fusions_amount + 1,
    #         5: worst_case_raw, 6: best_case_raw,
    #         7: naive_case, 8: worst_case, 9: best_case,
    #         10: naive_case / best_case, 11: worst_case / best_case,
    #         12: str(fp.error_angle), 13: str(fp.error_axis)}
    #
    # row_index = ws.max_row + 1
    # ws.insert_rows(row_index)
    # for col, value in data.items():
    #     ws.cell(row=row_index, column=col, value=value)
    #
    # book.save(filename=result_file_name)


if __name__ == '__main__':
    print("1 Init Fusion Gates...")

    error_angle = 0.1
    error_axis = [0, 1, 0]
    fm = FusionGate(error_angle=-error_angle, error_axis=error_axis)
    fp = FusionGate(error_angle=error_angle, error_axis=error_axis)
    fi = FusionGate(error_angle=0, error_axis=error_axis)



    # # print("O-O ~ O-O")
    # # c12p = Cluster.fuseLinearClusters(size_of_cluster=2, clicks='00', fusions=[fp])
    # # print("O-O ~ O-O")
    # # c12m = Cluster.fuseLinearClusters(size_of_cluster=2, clicks='00', fusions=[fm])
    # # c12i = Cluster.fuseLinearClusters(size_of_cluster=2, clicks='00', fusions=[fi])
    # # print("O-O ~ O-O")
    # # c12naive = Cluster.fuseLinearClusters(size_of_cluster=3, clicks='00', fusions=[fp]*4)
    # # c12pm = Cluster.fuseLinearClusters(size_of_cluster=3, clicks='00', fusions=[fp, fm]*2)
    # # c12i = Cluster.fuseLinearClusters(size_of_cluster=3, clicks='00', fusions=[fi]*4)
    # # print("PM dist from ideal : " + str(c12pm - c12i))
    # # print("Naive dist from ideal : " + str(c12naive - c12i))
    #
    for clicks in ["00"]:
        for mini_c_size in [3]:
            for fusions_amount in range(2, 3):
                runOverPMCombinations(fi, fp, fm, cluster_size=mini_c_size,
                                      clicks=clicks,
                                      fusions_amount=fusions_amount,
                                      rot_state_angle=0,
                                      rot_state_axis=[1, 0, 0])

                c12i = Cluster.fuseLinearClusters(size_of_cluster=cluster_size,
                                                  clicks=clicks,
                                                  fusions=[fi] * fusions_amount,
                                                  rot_state_angle=rot_state_angle,
                                                  rot_state_axis=rot_state_axis)
    #
    # for clicks in ["00", "01", "10", "11"]:
    #     for mini_c_size in [3]:
    #         for fusions_amount in range(2, 5):
    #             runOverPMCombinations(fi, fp, fm, cluster_size=mini_c_size,
    #                                   clicks=clicks,
    #                                   fusions_amount=fusions_amount,
    #                                   rot_state_angle=0,
    #                                   rot_state_axis=[1, 0, 0])
    #
    # for clicks in ["00", "01", "10", "11"]:
    #     for mini_c_size in [4]:
    #         for fusions_amount in range(2, 4):
    #             runOverPMCombinations(fi, fp, fm, cluster_size=mini_c_size,
    #                                   clicks=clicks,
    #                                   fusions_amount=fusions_amount,
    #                                   rot_state_angle=0,
    #                                   rot_state_axis=[1, 0, 0])
    #
    # for clicks in ["00", "01", "10", "11"]:
    #     for mini_c_size in [5]:
    #         for fusions_amount in range(2, 3):
    #             runOverPMCombinations(fi, fp, fm, cluster_size=mini_c_size,
    #                                   clicks=clicks,
    #                                   fusions_amount=fusions_amount,
    #                                   rot_state_angle=0,
    #                                   rot_state_axis=[1, 0, 0])



